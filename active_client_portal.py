import streamlit as st
import pandas as pd
import os
from openpyxl import load_workbook
from openpyxl.utils.exceptions import InvalidFileException
from datetime import datetime, timedelta
import plotly.express as px
import numpy as np
import smtplib
from email.mime.text import MIMEText
from email.mime.multipart import MIMEMultipart
import random
import string
from io import BytesIO
import requests

# Correct usage
bool_type = np.bool_

# Set most recently fully reviewed month-end close time period
today = datetime.today()
first_day_current_month = today.replace(day=1)
previous_month = first_day_current_month - timedelta(days=1)
first_day_next_month = (today + timedelta(days=30)).replace(day=1)
first_day_of_following_month = (today + timedelta(days=60)).replace(day=1)
next_month = first_day_of_following_month - timedelta(days=1)


# Set up dictionary for valid clients (Client ID: Password)
valid_clients = {
    "EI": "EI2024!",
    "AL": "AL2024!",
    "DLI": "DLI2024!",
    "DMSF": "DMSF2024!",
    "IA": "IA2024!",
    "LB": "LB2024!"
}
valid_client_names = {
    "EI": "Est Institute",
    "AL": "A&L Home Builders",
    "DLI": "Legacy Tattoo",
    "DMSF": "Darnerien McCants Sports & Fitness",
    "IA": "Intentionally Amazing",
    "LB": "La Bete LLC"
}
valid_client_emails = {
    "EI": "jordanlee2017@gmail.com",
    "AL": "jordanlee2017@gmail.com",
    "DLI": "jordanlee2017@gmail.com",
    "DMSF": "jordanlee2017@gmail.com",
    "IA": "jordanlee2017@gmail.com",
    "LB": "jordanlee2017@gmail.com"
}

# Assign each client with an industry based on business type
valid_client_business_type = {
    "EI": "Education Institution",
    "AL": "Home Builder",
    "DLI": "Tattoo Parlor",
    "DMSF": "Fitness Trainer",
    "IA": "Nail Salon",
    "LB": "Barbershop"
}

# Set up dictionary for industry key performance indicators (Industry: ["KPI #1", "KPI #2", "etc."])
industry_index = {
    "Other Services": ["Barbershop", "Nail Salon", "Tattoo Parlor"],
    "Construction": ["Home Builder", "Contracting Services"],
    "Educational Services": ["Education Institution", "Tutor Services"],
    "Fitness": ["Fitness Trainer"]
}

kpi_index = {
    "Other Services": ["# of Successful Appointments", "# of Active Clients", "# of Recurring Client Base", "# of Anticipated Appointments", "# of Anticipated Clients", "Appt Multiplier",
    "Realized / Effective Bill Rate", "Productivity Utilization", "Implied Tenure", "Average Revenue", "MRR", "LTV/CAC"],
    "Construction": ["Total Contract Bookings"],
    "Educational Services": ["# of Students", "Net New Students", "# of Sessions", "Total Billed Hours", "Realized / Effective Bill Rate", "Implied Tenure", "Average Revenue", "LTV/CAC"],
    "Fitness": ["# of Bookings", "# of Recurring Clients", "Total Bill Hours", "Realized / Effective Bill Rate", "Average Revenue/Client", "MRR", "LTC/CAC"]
}

# Function to generate a temporary password
def generate_temp_password(length=5):
    chars = string.ascii_letters + string.digits #+ string.punctuation
    return ''.join(random.choice(chars) for _ in range(length))

# Function to send the email
def send_email(recipient_email, temp_password):
    sender_email = "jmmgroupva@gmail.com"  # Replace with your email
    sender_password = "ltiw leaq pbxq pqqu"        # Replace with your email password or app password

    subject = "JMM Client Portal Temporary Password"
    body = f"Hello,\n\nYour temporary password is: {temp_password}\n\nPlease reach out to jmmgroupva@gmail.com if there are any issues."

    # Create email message
    msg = MIMEMultipart()
    msg['From'] = sender_email
    msg['To'] = recipient_email
    msg['Subject'] = subject

    # Attach the body text
    msg.attach(MIMEText(body, 'plain'))

    try:
        # Connect to the server
        with smtplib.SMTP("smtp.gmail.com", 587) as server:  # Use the appropriate SMTP server and port
            server.starttls()  # Upgrade the connection to secure
            server.login(sender_email, sender_password)  # Login
            server.send_message(msg)  # Send the email
            print("Email sent successfully!")

    except Exception as e:
        print(f"Error sending email: {e}")

# Streamlit Sidebar for Authentication

st.sidebar.title("Client Authentication")

# Phase 1: Collect Client ID
client_id = st.sidebar.text_input("Client ID", "")

if st.sidebar.button("Request Security Code"):
    if client_id in valid_clients:
        temp_password = generate_temp_password()
        st.session_state['temp_password'] = temp_password
        send_email(valid_client_emails[client_id], temp_password)
        st.sidebar.success("Temporary password sent to your email!")
    else:
        st.sidebar.error("Client ID not found.")

# Phase 2: Validate Passwords
if 'temp_password' in st.session_state:
    client_password = st.sidebar.text_input("Client Password", "AL2024!", type="password")
    encrypted_password = st.sidebar.text_input("Encrypted Passowrd (Sent to "f"{valid_client_emails[client_id]})", "test", type="password")

    if st.sidebar.button("Submit"):
        if client_id not in valid_clients:
            st.sidebar.error("Client ID not found.")
        elif valid_clients[client_id] != client_password:
            st.sidebar.error("Incorrect client password.")
        elif st.session_state['temp_password'] != encrypted_password:
            st.sidebar.error("Incorrect temporary password.")
        else:
            st.sidebar.success(f"{client_id} authenticated successfully!")
            st.session_state['authenticated'] = True

# Post-authentication content
if st.session_state.get('authenticated'):
    st.title(f"Welcome, {valid_client_names[client_id]}!")
    # Proceed with additional actions like loading data, etc,
    # Search for the financial forecast model using the Client ID and password
    folder_path = os.path.join(os.getcwd())  # Replace with actual folder path
    file_name = f"{client_id}_FFM.xlsx"
    file_path = os.path.join(folder_path, file_name)

    if os.path.exists(file_path):
        try:            
            workbook = load_workbook(filename=file_path, data_only=True, read_only=True, keep_vba=True)
            new_workbook = pd.ExcelFile(file_path)
            
            # st.success(f"Successfully opened {file_name}")
            st.subheader("Earnings Overview")

            # Load and display client-specific KPIs based on chosen industry above
            client_kpis = []
            for key, item in industry_index.items():
                for kpi in item:
                    # print(kpi)
                    if kpi == valid_client_business_type[client_id]:
                        client_kpis = kpi_index[key]
                        #print(client_kpis)
                    else:
                        continue
                else:
                    continue
            
            # Display the available sheets in the Excel file
            # sheet_names = workbook.sheetnames
            # st.write(f"Available sheets: {sheet_names}")
        
            # Let the user select a sheet to view
            # selected_sheet = st.sidebar.selectbox("Select a sheet to view:", sheet_names)
            df = new_workbook.parse("Monthly Detail")
            # Remove the first column from the header
            df = df.iloc[:, :-1].set_axis(df.columns[1:], axis=1)
            
            ws = workbook['Monthly Detail']
            
            # Create a list of lists to store the data
            data = []

            # Iterate through rows and append data to the list
            for row in ws.iter_rows(values_only=True):
                data.append(row)

            # Create a DataFrame, optionally setting the first row as headers
            df2 = pd.DataFrame(data)
            print(df2.iloc[:15, 38:42])
            # df2.columns = df.iloc[0]  # Set the first row as header
            df2 = df[1:]  # Remove the first row from the data

            # Reset index
            df2.reset_index(drop=True, inplace=True)
            df2.dropna(inplace=True)
            # print(df.columns)

            st.sidebar.subheader("Set date range to review")
            selected_review_start_date = st.sidebar.date_input("Select a start date to review:", value=datetime(previous_month.year, 1, 1))
            selected_review_end_date = st.sidebar.date_input("Select a end date to review:", value=datetime(2024, 12, 31)) #previous_month)
       

            if selected_review_start_date.month > 9:
                review_start_date = selected_review_start_date.strftime("%Y.%m")
            else:
                review_start_date = str(selected_review_start_date.year) + '.' + str(selected_review_start_date.month).strip('0')
            if selected_review_end_date.month > 9:
                review_end_date = selected_review_end_date.strftime("%Y.%m")
            else:
                review_end_date = str(selected_review_end_date.year) + '.' + str(selected_review_end_date.month).strip('0')

            formatted_start_date = selected_review_start_date.strftime('%B %Y')
            formatted_end_date = selected_review_end_date.strftime("%B %Y")

            # Create and add review date range
            review_cols = []
            tok = -1

            for col in df.columns:
                #print(col)
                if col == "Unnamed: 1" or col == "Unnamed: 2" or col == 2022:
                    continue
                elif col == 2023 or col == 2024 or col == 2025 or col == 2026 or col == 2027 or col == 2028 or col == 2029 or col == 2030:
                    new_col = col + tok
                    # tok += 1

                    new_date = str(new_col) + ".12"
                    df.rename(columns={col: new_date}, inplace=True)
                    col = new_date
                    if datetime.strptime(review_start_date, "%Y.%m") <= datetime.strptime(col, "%Y.%m") <= datetime.strptime(review_end_date, "%Y.%m"):
                        review_cols.append(col)
                    else:
                        continue
                    # review_cols.append(col)
                elif datetime.strptime(review_start_date, "%Y.%m") <= datetime.strptime(col, "%Y.%m") <= datetime.strptime(review_end_date, "%Y.%m"):
                    review_cols.append(col)
                else:
                    col = datetime.strptime(str(col), "%Y.%m")

            income_row = df.loc[df['Unnamed: 2'] == "Total Income"]
            income_row.set_index('Unnamed: 2', inplace=True)
            gm_row = df.loc[df["Unnamed: 2"] == 'Gross Profit']
            gm_row.set_index('Unnamed: 2', inplace=True)
            noi_row = df.loc[df["Unnamed: 2"] == 'Net Operating Income']
            noi_row.set_index('Unnamed: 2', inplace=True)
            ni_row = df.loc[df['Unnamed: 2'] == 'Net Income']
            ni_row.set_index('Unnamed: 2', inplace=True)

            cffo_row = df.loc[df['Unnamed: 2'] == 'Cash Flow From Operations']
            cffo_row.set_index('Unnamed: 2', inplace=True)
            cffi_row = df.loc[df['Unnamed: 2'] == 'Cash Flow From Investing']
            cffi_row.set_index('Unnamed: 2', inplace=True)
            cfff_row = df.loc[df['Unnamed: 2'] == 'Cash Flow From Financing']
            cfff_row.set_index('Unnamed: 2', inplace=True)
            cash_row = df.loc[df['Unnamed: 2'] == 'Ending Balance']
            cash_row.set_index('Unnamed: 2', inplace=True)

            mrr_row = df.loc[df[2022] == 'MRR']
            mrr_row.set_index(2022, inplace=True)
            
            formatted_cols = []
            # Create the filtered earnings DataFrame
            for col in review_cols:
                col = datetime.strptime(str(col), "%Y.%m")
                col = col.strftime("%B %Y")
                formatted_cols.append(col)

            earnings_df = pd.concat(
                [income_row, gm_row, noi_row],
                axis=0
            )[review_cols]
            
            earnings_df.columns = formatted_cols
            earnings_df.index.name = "Legend"
            

            st.write(f"Filtered Data from {formatted_start_date} to {formatted_end_date}:")

            # Plot the data
            if not earnings_df.empty:
                stacked_data = earnings_df.transpose().reset_index()
                stacked_data = stacked_data.melt(id_vars="index", var_name="Legend", value_name="Amount")
                earnings_df = earnings_df.style.format("${:,.2f}")
                st.dataframe(earnings_df)
                fig = px.bar(
                    stacked_data,
                    x="index",
                    y="Amount",
                    color="Legend",
                    title="Income, Gross Profit, and NOI (Filtered)",
                    labels={"index": "Date", "Amount": "Amount ($)"},
                    barmode="stack"
                )
                st.plotly_chart(fig)
            else:
                st.warning("No data available for the selected date range.")

        
            st.subheader("Cash Position Overview")

            # Create the filtered cash position DataFrame
            cash_df = pd.concat(
                [mrr_row, cffo_row, cffi_row, cfff_row, cash_row],
                axis=0
            )[review_cols]
            
            cash_df.columns = formatted_cols
            cash_df.index.name = "Legend"
            
            st.write(f"Filtered Data from {formatted_start_date} to {formatted_end_date}:")
            import plotly.graph_objects as go
            # Plot the data
            if not cash_df.empty:
                stacked_data = cash_df.transpose().reset_index()
                stacked_data = stacked_data.melt(id_vars="index", var_name="Legend", value_name="Amount")
                styled_cash_df = cash_df.style.format("${:,.2f}")
                st.dataframe(styled_cash_df)
                fig = go.Figure()
                # print(cash_df.loc["Cash Flow From Investing"])

                # Adding bar chart for CFFO, CFFI, CFFF, and Cash Ending Balance
                for metric in ["Cash Flow From Operations", "Cash Flow From Investing", "Cash Flow From Financing", "Ending Balance"]:
                    if metric in cash_df.index:
                        fig.add_trace(go.Bar(
                            x=cash_df.columns,
                            y=cash_df.loc[metric],
                            name=metric
                        ))
                    else:
                        st.warning(f"{metric} not found in cash data. Check spelling or index structure.")
                if "MRR" in client_kpis:
                        

                    # Adding MRR as a line chart with a secondary y-axis
                    fig.add_trace(go.Scatter(
                        x=cash_df.columns,
                        y=cash_df.loc["MRR"],
                        mode='lines+markers',
                        name='MRR',
                        yaxis='y2'
                    ))
                    # Update layout with secondary y-axis for MRR
                    fig.update_layout(
                        title="Cash Flow and MRR Overview (Filtered)",
                        xaxis=dict(
                        title='Date',
                        tickangle=45  # Rotate x-axis labels for readability
                        ),
                        
                        yaxis=dict(
                            title='Amount ($)',
                            automargin=True  # Adjusts margins automatically for better spacing
                        ),
                        yaxis2=dict(
                            title="MRR ($)",
                            overlaying='y',  # Overlay with primary y-axis
                            side='right',  # Position on the right side
                            automargin=True
                        ),
                        legend=dict(
                            x=0.5, y=1.15,  # Adjust legend position for better visibility
                            xanchor='center',
                            orientation='h'  # Horizontal legend
                        ),
                        margin=dict(
                            l=80, r=80, t=100, b=100  # Increase margins to ensure no clipping
                        ),
                        barmode='stack',  # Stack bars for cash flow data
                        plot_bgcolor='white'  # Set background color for better contrast
                    )

                    # Display the chart
                    st.plotly_chart(fig)
                else:
                    fig.update_layout(
                        title="Cash Flow (Filtered)",
                        xaxis=dict(title='Date'),
                        yaxis=dict(title='Amount ($)'),
                        barmode='stack'
                    )
                    st.plotly_chart(fig)

                
            else:
                st.warning("No data available for the selected date range.")

            st.subheader("Key Performance Indicator Overview")

            kpi_dfs = []
            for kpi in client_kpis:
                if kpi != "MRR":
                    kpi_row = df.loc[df[2022] == kpi]
                    kpi_row.set_index(2022, inplace=True)
                    kpi_dfs.append(kpi_row)
                else:
                    continue
            #Sprint(kpi_dfs)

            # Create the filtered KPIs DataFrame
            kpi_df = pd.concat(
                kpi_dfs,
                axis=0
            )[review_cols]
            #print(kpi_df)

            st.sidebar.subheader("Set date range to adjust the metrics below")
            selected_adjusted_start_date = st.sidebar.date_input("Select the start date of the date range to adjust:", value=first_day_next_month)
            selected_adjusted_end_date = st.sidebar.date_input("Select the end date of the date range to adjust:", value=next_month)

            st.sidebar.subheader("Key Performance Indicators")

            kpi_toggles = []
            for i in client_kpis:
                if i != "MRR":
                    kpi_toggle = st.sidebar.number_input(i, kpi_df.loc[i, review_end_date])
                    kpi_toggles.append(kpi_toggle)
                else:
                    continue

            def has_sidebar_number_input_changed(current_value) -> bool:
                """
                Checks if a user changed a `st.sidebar.number_input`.
                
                Args:
                    key (str): A unique key for the number_input widget.
                    label (str): The label displayed for the number_input widget.
                    default_value (float): The default value for the number_input widget.
                    step (float): The increment step for the number_input widget.
                
                Returns:
                    bool: True if the number_input value changed, False otherwise.
                """
                if 'previous_values' not in st.session_state:
                    st.session_state['previous_values'] = {}

                # Compare with the previous value
                previous_value = st.session_state['previous_values'].get(key)
                value_changed = current_value != previous_value

                # Update the previous value
                # st.session_state['previous_values'][key] = current_value

                return value_changed, current_value

            def adjust_forecast_kpi(dataframe, client_kpis, start_date, end_date):
                """
                Adjust KPI values in the forecast DataFrame based on user inputs and validate date ranges.

                Parameters:
                dataframe (pd.DataFrame): The forecast DataFrame with a 'date' column and KPI columns.
                kpi_name (str): The name of the KPI column to adjust.
                start_date (str): The start date of the adjustment range (YYYY-MM-DD).
                end_date (str): The end date of the adjustment range (YYYY-MM-DD).
                adjustment_percentage (float): Percentage change to apply (e.g., 10 for +10%).
                review_date_range (tuple): Tuple of start and end dates for the review range (YYYY-MM-DD).

                Returns:
                pd.DataFrame: Adjusted DataFrame.
                str: Warning message if date range exceeds the review range.
                """
                '''review_start = review_s
                review_end = review_e'''


                '''# Validate date ranges
                if start_date < review_start or end_date > review_end:
                    return st.warning("Warning: Adjustment date range exceeds the review date range.")
                '''
                for kpi_name in client_kpis:
                    if has_sidebar_number_input_changed(kpi_name)[0]:
                        # Apply adjustment
                        adjustment = has_sidebar_number_input_changed(kpi_name)[1]
                        print(dataframe)
                        mask = (dataframe['date'] >= start_date) & (dataframe['date'] <= end_date)
                        dataframe.loc[mask, kpi_name] = adjustment
                    else:
                        continue

                return dataframe

            if st.sidebar.button("Apply Adjustment"):
                kpi_df = adjust_forecast_kpi(kpi_df, client_kpis=client_kpis, start_date=selected_adjusted_start_date, 
                end_date=selected_adjusted_end_date)
            

            kpi_df.columns = formatted_cols
            kpi_df.index.name = "Legend"
            
            st.write(f"Filtered Data from {formatted_start_date} to {formatted_end_date}:")
            st.dataframe(kpi_df)

            # Plot the data
            if not kpi_df.empty:
                stacked_data = kpi_df.transpose().reset_index()
                stacked_data = stacked_data.melt(id_vars="index", var_name="Legend", value_name="Amount")
                 # Create the Plotly figure with a secondary y-axis
                fig = go.Figure()

                # Loop through KPI metrics and add them to the plot
                for metric in kpi_df.index:
                    if metric in kpi_df.index:
                        yaxis_type = 'y2' if kpi_df.loc[metric].max() < 10 else 'y'
                        fig.add_trace(go.Scatter(
                            x=stacked_data[stacked_data['Legend'] == metric]['index'],
                            y=stacked_data[stacked_data['Legend'] == metric]['Amount'],
                            mode='lines',
                            name=metric,
                            yaxis=yaxis_type
                        ))
                    else:
                        st.warning(f"{metric} not found in cash data. Check spelling or index structure.")

                fig.update_layout(
                    title="Key Performance Indicators (Filtered)",
                    xaxis=dict(title='Date'),
                    yaxis=dict(title='Amount'),
                    yaxis2=dict(
                        title='Low-value KPI Metrics',
                        overlaying='y',
                        side='right'
                    )
                )
                st.plotly_chart(fig)
            else:
                st.warning("No data available for the selected date range.")
            
            # Load the selected sheet into a DataFrame
            # active_sheet = workbook[selected_sheet]
            # data = active_sheet.values
            # sheet_data = pd.DataFrame(data)

            # Display the sheet data
            # st.dataframe(sheet_data)
        except InvalidFileException:
            st.error("Unable to open the file. The file may be corrupt or inaccessible.")
    else:
        st.error(f"No financial forecast model found for Client ID '{client_id}'.")

else:
    st.write("Please enter your credentials to proceed.")
