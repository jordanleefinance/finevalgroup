import openpyxl
from openpyxl.utils import get_column_letter
import subprocess
import time
#from formulas import ExcelCompiler
class ForecastUpdater:
    def __init__(self, file_path):
        self.unprotected_file_path = file_path
    def update_forecast_to_values(self, file_path):

        # Step 1: Open file in Excel to recalculate and cache formulas
        print("Opening Excel to recalculate formulas...")
        subprocess.run(['start', 'excel', file_path], shell=True)
        time.sleep(5)  # Wait for Excel to open
        subprocess.run(['taskkill', '/IM', 'excel.exe', '/F'], shell=True)  # Close Excel
        time.sleep(2)

        '''# Compile and evaluate Excel formulas
        xl = ExcelCompiler(excel=file_path)
        xl.compile()'''

        wb = openpyxl.load_workbook(file_path)
        try:
            ws = wb['Budget to Actual']
        except KeyError:
            ws = wb['Actual vs. Forecast']

        # Step 2: Find the row containing 'Forecast', 'Actual', and 'Budget'
        forecast_col = None
        actual_col = None
        budget_col = None

        for row in ws.iter_rows():
            for cell in row:
                if cell.value == 'Forecast':
                    forecast_col = cell.column
                if cell.value == 'Actual':
                    actual_col = cell.column
                if cell.value == 'Budget':
                    budget_col = cell.column
            if forecast_col and actual_col and budget_col:
                target_row = cell.row
                break
        print(f"Found: Forecast={forecast_col}, Actual={actual_col}, Budget={budget_col}, Target Row={target_row}")

        # Step 3: Insert a column after 'Actual' for 'Actuals'
        wb_data_only = openpyxl.load_workbook(file_path, data_only=True)
        try:
            ws_data_only = wb['Budget to Actual']
        except KeyError:
            ws_data_only = wb['Actual vs. Forecast']
        # Convert the NEW Forecast column to values (replace formulas with evaluated values)
        # find original Forecast column index in the data-only workbook (header row = target_row)
        orig_forecast_col = None
        for c in range(1, ws_data_only.max_column + 1):
            if ws_data_only.cell(row=target_row, column=c).value == 'Actual':
                orig_forecast_col = c + 1  # Original Forecast is 2 columns after Actual
                print(f"Found original Forecast col in data-only sheet: {orig_forecast_col}")
        if orig_forecast_col is None:
            orig_forecast_col = actual_col # fallback

        # Overwrite new_forecast_col with values from the data-only sheet (skip header)
        for r in range(target_row + 1, ws.max_row + 1):
            val = ws_data_only.cell(row=r, column=orig_forecast_col-1).value
            ws.cell(row=r, column=orig_forecast_col-1).value = val
        wb_data_only.close()
        wb.save(file_path)
        wb.close()

# Usage example
if __name__ == "__main__":

    # Path to the original Excel file
    original_file_path = r'C:\Users\jorda\OneDrive\Documents\GitHub\finevalgroup\IA_FFM_2025_B2A.xlsx'
    updater = ForecastUpdater(original_file_path)
    updater.update_forecast_to_values(original_file_path)
