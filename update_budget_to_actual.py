import openpyxl
from openpyxl.utils import get_column_letter
import os
from copy import copy
from datetime import datetime
from datetime import datetime, timedelta

class BudgetToActualUpdater:
    def __init__(self, file_path, new_file_path = None, close_month=None):
        self.file_path = file_path
        self.close_month = close_month
        if new_file_path is None:
            self.new_file_path = self._generate_new_file_path()
        else:
            self.new_file_path = new_file_path

    def _generate_new_file_path(self):
        base, ext = os.path.splitext(self.file_path)
        return f"{base}_Updated{ext}"
    
    def update_budget_to_actual(self, file_path, new_file_path = None, close_month=datetime.today()):
        # Load the workbook and select the 'Budget to Actual' tab
        #file_path = 'SandBox_FFM.xlsx'  # Update this path to your file location
        wb = openpyxl.load_workbook(file_path)
        try:
            ws = wb['Budget to Actual']
        except KeyError:
            ws = wb['Actual vs. Forecast']
            

        # Ungroup all columns initially
        for col_letter in list(ws.column_dimensions.keys()):
            try:
                ws.column_dimensions.ungroup(col_letter, col_letter)
            except Exception:
                pass

        date_header_range = None
        # 1) Unmerge any merged ranges that include the first row (date row)
        for merged in list(ws.merged_cells.ranges):
            if merged.min_row == 14:  # Assuming the date header is in row 14
                # Store the value of the merged cell
                merged_value = ws.cell(row=merged.min_row, column=merged.min_col).value
                # Unmerge the cells
                ws.unmerge_cells(str(merged))
                # Set the value back to the top-left cell of the unmerged range
                ws.cell(row=merged.min_row, column=merged.min_col, value=merged_value)
                date_header_range = merged
                print(merged)
                
        # Find the last date of the previous month
        today = datetime.today()
        if close_month is None:
            close_month = today.replace(day=1) - timedelta(days=1)
        first_of_current_month = close_month.replace(day=1)
        last_of_previous_month = first_of_current_month - timedelta(days=1)

        # Step 2: Find the row containing 'Forecast', 'Actual', and 'Budget'
        forecast_col = None
        actual_col = None
        budget_col = None

        for row in ws.iter_rows():
            for cell in row:
                if cell.value == 'Forecast':
                    forecast_col = cell.column
                if cell.value == 'Actual':
                    actual_col = cell.column
                if cell.value == 'Budget':
                    budget_col = cell.column
            if forecast_col and actual_col and budget_col:
                target_row = cell.row
                break
        print(f"Found: Forecast={forecast_col}, Actual={actual_col}, Budget={budget_col}, Target Row={target_row}")

        # Step 3: Insert a column to the right of 'Forecast' and 'Actual'
        ws.insert_cols(forecast_col + 1)
        new_forecast_col = forecast_col + 1  # New column index for 'Forecasts'
        ws.insert_cols(actual_col + 2)
        new_actuals_col = actual_col + 2  # New column index for 'Actuals'

        wb.save(new_file_path)

        wb = openpyxl.load_workbook(new_file_path)
        try:
            ws = wb['Budget to Actual']
        except KeyError:
            ws = wb['Actual vs. Forecast']
        #actual_col += 1  # Adjust actual_col due to previous insertion
        #ws.insert_cols(actual_col + 2)  # Accounting for the shift due to the previous insertion
        print(f"After insert: Forecast col={forecast_col}, New Forecast col={new_forecast_col}, Actual col={actual_col}, New Actual col={new_actuals_col}")

        # Step 4: Copy formulas and ALL formatting from the 'Actual' column to the new column
        

        for orginal_cell, new_cell in zip(ws[ get_column_letter(actual_col+1) ], ws[ get_column_letter(new_actuals_col) ]):
            # Copy value/formula
            if orginal_cell.data_type == 'f':  # If it's a formula
                new_cell.value = orginal_cell.value  # Copy formula as-is
            else:
                new_cell.value = orginal_cell.value  # Copy value
            
            # Copy ALL formatting
            try:
                new_cell.font = copy(orginal_cell.font)
                # Copy fill (colors)
                if orginal_cell.fill:
                    new_cell.fill = copy(orginal_cell.fill)
                new_cell.border = copy(orginal_cell.border)
                new_cell.alignment = copy(orginal_cell.alignment)
                new_cell.number_format = orginal_cell.number_format
                new_cell.protection = copy(orginal_cell.protection)
                # Copy comment if present
                if orginal_cell.comment:
                    new_cell.comment = copy(orginal_cell.comment)
                
            except Exception as e:
                print(f"Warning: Could not copy formatting for row {orginal_cell.row}: {e}")
                pass
        # Step 4b: Convert the original 'Actual' column to values only (remove formulas)
        wb_data_only = openpyxl.load_workbook(file_path, data_only=True)
        try:
            ws_data_only = wb['Budget to Actual']
        except KeyError:
            ws_data_only = wb['Actual vs. Forecast']
        for orginal_cell, new_cell in zip(ws[ get_column_letter(actual_col) ], ws[ get_column_letter(new_actuals_col) ]):
            row = orginal_cell.row
            if orginal_cell.row == target_row:
                original_cell_value = ws.cell(row=target_row, column=actual_col+1, value=last_of_previous_month.strftime('%#m/%d/%Y')).value
            else:
                original_cell_value = ws_data_only.cell(row=row, column=actual_col).value
            ws.cell(row=row, column=actual_col+1).value = original_cell_value
        wb_data_only.close()

        # Step 5: Update the top row (with dates) with the last date of the previous month
        # Use the date_header_range to determine the header row
        if date_header_range:
            header_row_for_date = date_header_range.min_row
            header_col_for_date = date_header_range.min_col
            print(f"Updating date in row {header_row_for_date}, column {header_col_for_date}")
            print(f"Setting date in row {header_row_for_date}, column {new_actuals_col}")
            ws.cell(row=header_row_for_date, column=header_col_for_date, value=close_month)  # Example: setting to July 31, 2024
        else:
            # Fallback if no merged range found
            ws.cell(row=target_row, column=actual_col, value=close_month)

        # Step 6b: Convert the original 'Forecast' column to values only (remove formulas)
        for orginal_cell, new_cell in zip(ws[ get_column_letter(new_actuals_col) ], ws[ get_column_letter(new_forecast_col) ]):
            # Copy value/formula
            if orginal_cell.data_type == 'f':  # If it's a formula
                new_cell.value = orginal_cell.value  # Copy formula as-is
            else:
                new_cell.value = orginal_cell.value  # Copy value
            
            # Copy ALL formatting
            try:
                new_cell.font = copy(orginal_cell.font)
                # Copy fill (colors)
                if orginal_cell.fill:
                    new_cell.fill = copy(orginal_cell.fill)
                new_cell.border = copy(orginal_cell.border)
                new_cell.alignment = copy(orginal_cell.alignment)
                new_cell.number_format = orginal_cell.number_format
                new_cell.protection = copy(orginal_cell.protection)
                # Copy comment if present
                if orginal_cell.comment:
                    new_cell.comment = copy(orginal_cell.comment)
                
            except Exception as e:
                print(f"Warning: Could not copy formatting for row {orginal_cell.row}: {e}")
                pass

        for row in range(target_row, ws.max_row + 1):  # Start AFTER the header row
            if row == target_row:
                ws.cell(row=row, column=new_forecast_col).value = 'Forecast'
                ws.cell(row=row, column=new_forecast_col-1).value = last_of_previous_month.strftime('%#m/%d/%Y')


        for row in range(target_row, ws.max_row + 1):  # Start AFTER the header row
            if row == target_row:
                ws.cell(row=row, column=budget_col+3).value = 'Budget'
                ws.cell(row=row, column=budget_col+2).value = last_of_previous_month.strftime('%#m/%d/%Y')



        # Step 6: Rerun the merge if needed
        if date_header_range:
            ws.merge_cells(start_row=date_header_range.min_row, start_column=date_header_range.min_col,
                            end_row=date_header_range.max_row, end_column=date_header_range.max_col+2)
            
        # Step 7: Find and update 'Variance (Forecast)' and 'Variance (Budget)' columns
        variance_forecast_col = None
        variance_budget_col = None

        # Find the variance columns by searching the header row
        for col in range(1, ws.max_column + 1):
            cell_value = ws.cell(row=target_row, column=col).value
            if cell_value == 'Variance (Forecast)':
                variance_forecast_col = col
            elif cell_value == 'Variance (Budget)':
                variance_budget_col = col

        print(f"Found: Variance (Forecast) col={variance_forecast_col}, Variance (Budget) col={variance_budget_col}")

        # Update formulas in 'Variance (Forecast)' column
        # Formula should be: Actual - Forecast
        if variance_forecast_col:
            for row in range(target_row + 1, ws.max_row-1):
                # Create formula: new_actuals_col - new_forecast_col
                formula = f"={get_column_letter(new_actuals_col)}{row}-{get_column_letter(new_forecast_col)}{row}"
                ws.cell(row=row, column=variance_forecast_col).value = formula

        # Update formulas in 'Variance (Budget)' column
        # Formula should be: Actual - Budget
        if variance_budget_col:
            for row in range(target_row + 1, ws.max_row-1):
                # Create formula: new_actuals_col - budget_col
                formula = f"={get_column_letter(new_actuals_col)}{row}-{get_column_letter(budget_col+3)}{row}"
                ws.cell(row=row, column=variance_budget_col).value = formula

        # Step 8: Regroup columns, excluding specific ones


        # header row
        header_row = target_row if 'target_row' in globals() else 1

        # Toggle: if True, exclude these headers from grouping; set False to include them
        exclude_headers_from_grouping = True

        # Names to exclude when flag is True (normalized)
        default_exclude = {
            "forecast", "actual", "budget",
            "variance (forecast)", "variance (budget)"
        }

        # Build header map: col_index -> raw header value
        header_map = {}
        for col in range(1, ws.max_column + 1):
            header_map[col] = ws.cell(row=header_row, column=col).value

        # Debug output to help diagnose grouping issues
        print("Header row values:")
        '''for col, val in header_map.items():
            print(col, "->", repr(val))
        '''
        # Decide which names to exclude
        if exclude_headers_from_grouping:
            exclude_names = {n.lower() for n in default_exclude}
        else:
            exclude_names = set()

        # Identify date columns: real datetimes or strings that parse as dates
        date_columns = []
        for col, val in header_map.items():
            if val is None:
                continue
            if isinstance(val, datetime):
                date_columns.append(col)
                continue
            txt = str(val).strip()
            if txt.lower() in exclude_names:
                # explicitly excluded
                continue
            # try parsing common formats
            parsed = None
            for fmt in ("%m/%d/%Y", "%Y-%m-%d", "%B %Y", "%b %Y", "%m/%Y", "%d-%b-%Y"):
                try:
                    parsed = datetime.strptime(txt, fmt)
                    break
                except Exception:
                    pass
            if parsed:
                date_columns.append(col)

        date_columns = sorted(date_columns)
        print("Detected date columns:", date_columns)


        # Group only consecutive date columns
        if date_columns:
            start = date_columns[0]
            prev = start
            for c in date_columns[1:]:
                if c == 7:
                    continue  # Skip grouping column B (7)
                if c == prev + 1:
                    prev = c
                    continue
                ws.column_dimensions.group(get_column_letter(start), get_column_letter(prev), hidden=False)
                start = c
                prev = c
            ws.column_dimensions.group(get_column_letter(start), get_column_letter(prev), hidden=False)

        # Step 11: Ensure new columns are NOT grouped
        try:
            ws.column_dimensions[get_column_letter(new_actuals_col)].outline_level = 0
            ws.column_dimensions[get_column_letter(new_forecast_col)].outline_level = 0
        except Exception:
            pass

        # Save the workbook
        wb.save(new_file_path)
        wb.close()

        import os
        import time
        import streamlit as st
        #import msvcrt

        from copy_paste_forecast import ForecastUpdater
        updater = ForecastUpdater(new_file_path)

        # Open the file in Excel and wait for the user to save/close it (or press Enter).
        # This waits for either: a change to the file modification time, the user pressing Enter,
        # or the timeout (seconds) being reached.
        initial_mtime = os.path.getmtime(new_file_path)
        try:
            os.startfile(new_file_path)  # opens in default app (Excel on Windows)
        except Exception as e:
            with open(new_file_path, mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet") as f:
                st.download_button('Download updated file', f, file_name=os.path.basename(new_file_path))
                print(f"Please open the file manually: {new_file_path}")
        timeout = 180  # seconds, adjust as needed
        start = time.time()
        print(f"Opened {new_file_path}. Please save & close Excel, or press Enter here when done. Waiting up to {timeout}s...")

        while True:
            # User pressed a key?
            '''if msvcrt.kbhit():
                ch = msvcrt.getwch()
                if ch in ("\r", "\n"):
                    print("User confirmed via Enter. Continuing...")
                    break'''

            # File modification detected?
            try:
                if os.path.getmtime(new_file_path) != initial_mtime:
                    print("File modification detected. Continuing...")
                    break
            except Exception:
                pass

            # Timeout?
            if time.time() - start > timeout:
                print("Timeout reached. Continuing...")
                break

            time.sleep(0.5)

        # Now run the updater (safe to assume file has been saved)
        updater.update_forecast_to_values(new_file_path)
