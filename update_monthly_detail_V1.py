import openpyxl
import os
from datetime import datetime, timedelta
from openpyxl.utils import get_column_letter, column_index_from_string

class ExcelProcessor:
    def __init__(self, original_file_path, close_month=None):
        
        self.original_file_path = os.path.abspath(original_file_path)
        self.close_month = close_month
        self.unprotected_file_path = self._generate_unprotected_file_path()
        #print(self.original_file_path)

        # immediate sanity checks with helpful debug info
        if not os.path.exists(self.original_file_path):
            parent = os.path.dirname(self.original_file_path) or os.getcwd()
            raise FileNotFoundError(
                f"Original file not found: {self.original_file_path}\n"
                f"Current working dir: {os.getcwd()}\n"
                f"Parent dir: {parent}\n"
                f"Parent dir listing: {os.listdir(parent) if os.path.isdir(parent) else 'N/A'}"
            )

    def _generate_unprotected_file_path(self):
        base, ext = os.path.splitext(self.original_file_path)
        return f"{base}_Updated{ext}"
    
    def update_budget_to_actual(self, close_month=None):
        from update_budget_to_actual import BudgetToActualUpdater
        updater = BudgetToActualUpdater(self.original_file_path, new_file_path=self.unprotected_file_path, close_month=close_month)
        updater.update_budget_to_actual(self.original_file_path, new_file_path=self.unprotected_file_path, close_month=close_month)

    def remove_password(self):
        try:
            # Verify if the file is an Excel file
            if not self.original_file_path.endswith(('.xlsx', '.xlsm')):
                raise ValueError("The file is not a valid Excel file.")

            # choose keep_vba only for xlsm
            keep_vba = self.original_file_path.endswith('.xlsm')
            workbook = openpyxl.load_workbook(os.path.abspath(self.original_file_path), read_only=False, keep_vba=keep_vba)

            # Save a new copy without password protection
            if not os.path.exists(self.unprotected_file_path):
                workbook.save(self.unprotected_file_path)
                print(f"Password removed successfully. New file saved as: {self.unprotected_file_path}")
            else:
                print(f"Unprotected file already exists: {self.unprotected_file_path}")
        except ValueError as ve:
            print(f"An error occurred: {ve}")
        except Exception as e:
            print(f"An error occurred while removing password: {e}")

    def find_date_in_row(self, sheet_name='Monthly Detail', search_row=4, target_date=None):
        try:
            # Load the workbook and select the sheet
            workbook = openpyxl.load_workbook(self.unprotected_file_path, data_only=True)
            sheet = workbook[sheet_name]

            # Set the target date to the last day of the previous month if not provided
            if target_date is None:
                target_date = datetime.today().replace(day=1) - timedelta(days=1)

            # Search for the target date in the specified row
            for cell in sheet[search_row]:
                if isinstance(cell.value, datetime) and cell.value.date() == target_date.date():
                    column_index = cell.column
                    column_letter = get_column_letter(column_index)

                    # Safely calculate previous and next column letters
                    pre_pre_column_letter = get_column_letter(column_index - 2) if column_index > 2 else None
                    pre_column_letter = get_column_letter(column_index - 1) if column_index > 1 else None
                    post_column_letter = get_column_letter(column_index + 1)

                    print(f"Found date {target_date} in column {column_letter}.")
                    return (
                        f"{pre_pre_column_letter}:{pre_pre_column_letter}" if pre_pre_column_letter else None,
                        f"{pre_column_letter}:{pre_column_letter}" if pre_column_letter else None,
                        f"{column_letter}:{column_letter}",
                        f"{post_column_letter}:{post_column_letter}",
                    )

            print(f"Date {target_date.date()} not found in row {search_row}.")
            return None
        except Exception as e:
            print(f"An error occurred while finding the date: {e}")
            return None

    def copy_formatting_and_formulas(self, sheet_name='Monthly Detail', target_date=None):
        try:
            # Find the source and target columns
            result = self.find_date_in_row(sheet_name, target_date=target_date)
            if result is None:
                raise ValueError("Date not found in the specified row.")

            pre_pre_source, pre_source, source, target = result

            # Validate column ranges
            if not source or not target:
                raise ValueError("Source or target column is invalid.")

            # Load the workbook and select the sheet
            workbook = openpyxl.load_workbook(self.unprotected_file_path)
            sheet = workbook[sheet_name]
            from copy import copy
            '''data_start_row = 5  # Assuming data starts from row 5
            max_row = sheet.max_row'''

                
            for source_cell, target_cell in zip(sheet[source], sheet[target]):
                if source_cell.value == target_date:
                    continue  # Skip copying if the target cell is the date header
                else:
                    target_cell.value = source_cell.value
                    target_cell.number_format = source_cell.number_format
                    target_cell.font = copy(source_cell.font)
                    target_cell.fill = copy(source_cell.fill)
                    target_cell.border = copy(source_cell.border)
                    target_cell.alignment = copy(source_cell.alignment)


            # Copy formatting and formulas
            if pre_source:
                for pre_source_cell, source_cell in zip(sheet[pre_source], sheet[source]):
                    if source_cell.value == target_date:
                        continue  # Skip copying if the target cell is the date header
                    else:
                        source_cell.value = pre_source_cell.value
                        source_cell.number_format = pre_source_cell.number_format
                        source_cell.font = copy(pre_source_cell.font)
                        source_cell.fill = copy(pre_source_cell.fill)
                        source_cell.border = copy(pre_source_cell.border)
                        source_cell.alignment = copy(pre_source_cell.alignment)

            if pre_pre_source:
                for pre_pre_source_cell, pre_source_cell in zip(sheet[pre_pre_source], sheet[pre_source]):
                    if pre_source_cell.value == target_date:
                        continue  # Skip copying if the target cell is the date header
                    else:
                        # pre_source_cell.value = pre_pre_source_cell.value
                        pre_source_cell.number_format = pre_pre_source_cell.number_format
                        pre_source_cell.font = copy(pre_pre_source_cell.font)
                        #pre_source_cell.fill = copy(pre_pre_source_cell.fill)
                        pre_source_cell.border = copy(pre_pre_source_cell.border)
                        pre_source_cell.alignment = copy(pre_pre_source_cell.alignment)

            

            # Save changes
            workbook.save(self.unprotected_file_path)
            print("Formatting and formulas copied successfully.")
        except Exception as e:
            print(f"An error occurred while copying formatting: {e}")

# Usage example
if __name__ == "__main__":
    month = datetime(2024, 8, 31)

    # Path to the original Excel file
    original_file_path = r'C:\Users\jorda\OneDrive\Documents\GitHub\finevalgroup\EI_FFM_2025.xlsx'

    # Initialize the processor
    processor = ExcelProcessor(original_file_path)
    # Update budget to actual
    processor.update_budget_to_actual(close_month=month)
    # Remove password protection
    processor.remove_password()

    # Copy formatting and formulas
    processor.copy_formatting_and_formulas(target_date=month)
